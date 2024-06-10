from flask import Flask, jsonify, request
from flask_marshmallow import Marshmallow
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook #load_workbook func is used to load .xlsx file into our python prog. to read data from or write data to the workbook
from flask_jwt_extended import (
    create_access_token,
    jwt_required,
    JWTManager,
    get_jwt_identity
)

app = Flask(__name__)

app.config["SQLALCHEMY_DATABASE_URI"] = (
    "sqlite:///user.db"  # it configures database connections..
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"]="thisissecret"

db = SQLAlchemy(app)
ma = Marshmallow(app)
api = JWTManager(app)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, nullable=False)
    username = db.Column(db.String(20), nullable=False)
    email = db.Column(db.String(50), nullable=False)
    password = db.Column(db.String(50), nullable=False)


class Userschema(ma.Schema): #Studentschema class ka use kisi student object ko JSON format mein serialize karne ke liye hota hai.
    class Meta:
        fields = ["user_id", "username", "email","password"] #Is line se JSON serialization ke liye chune gaye fields ka list specify kiya gaya hai.


user_schema = Userschema() #student_schema ek instance hai of Studentschema class.Iska use kisi single student object ko serialize karne ke liye hota hai.
user_schemas = Userschema(many=True) #Iska use multiple students ki list ko serialize karne ke liye hota hai.


#  adding data in db...
@app.route("/add", methods=["POST"])
def add_data():

    user_data = request.files["users"]#Yeh line POST request ke saath bheje gaye file ko retrieve karta hai. File ka name "Student" hai.
    Myuser = load_workbook(user_data)
    Newuser = Myuser.active

    for i in Newuser.iter_rows(min_row=2, values_only=True):
        if None not in i:
            user = User(user_id=i[0],username=i[1], email=i[2],password=i[3])
            db.session.add(user)
            db.session.commit()
    return "Message : Data retrieved..."

@app.route("/login", methods=["POST"])
def login():
    data = request.get_json()
    username = data.get("username")
    password = data.get("password")
    user = User.query.filter_by(username=username).first()
    if user and user.password==password:
        token = create_access_token(identity=user.id)
        return {"access token": token}
    else:
        return {"Message": "Invalid Login Credentials."}


#  Get all student
@app.route("/get", methods=["GET"])
@jwt_required()
def fetch_data():
    all_posts = User.query.all()  # query all data from db...
    result = user_schemas.dump(all_posts)
    return jsonify(result)


#  Get student by individual id
@app.route("/get/<int:id>", methods=["GET"])
@jwt_required()
def get_data(id):
    post = User.query.get(id)
    result = user_schema.dump(post)
    return jsonify(result)


#  update student in db....
@app.route("/update/<int:id>", methods=["PUT"])
@jwt_required()
def update_user(id):
    post = User.query.get(id)
    user_id = request.json.get("user_id")
    name = request.json.get("name")
    email = request.json.get("email")
    password = request.json.get("password")
    if name:
        post.name = name
    if email:
        post.email = email
    if password:
        post.password = password
    # db.session.add(post)
    db.session.commit()
    return user_schema.dump(post)


# delete book from db...
@app.route("/delete/<int:id>", methods=["DELETE"])
@jwt_required()
def delete_user(id):
    delete = User.query.get(id)
    db.session.delete(delete)
    db.session.commit()
    return user_schema.jsonify(delete)


if __name__ == "__main__":
    app.run(debug=True)